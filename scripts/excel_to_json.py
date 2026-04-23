import json
from pathlib import Path
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent.parent
INPUT_FILES = {
    "מפקח": BASE_DIR / "מפקח.xlsx",
    "משפטנים": BASE_DIR / "משפטנים.xlsx",
    "נגדים": BASE_DIR / "נגדים.xlsx",
    "פקד": BASE_DIR / "פקד.xlsx",
    "רפק": BASE_DIR / "רפק.xlsx",
}
OUTPUT_DIR = BASE_DIR / "public" / "data"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

def clean_value(value):
    if value is None:
        return None
    if isinstance(value, float):
        return round(value, 2)
    return value

def extract_sheet_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_row_index = 0
    for i, row in enumerate(rows[:10]):
        non_empty = [cell for cell in row if cell is not None]
        if len(non_empty) >= 3:
            header_row_index = i
            break

    headers = []
    for idx, cell in enumerate(rows[header_row_index]):
        headers.append(str(cell).strip() if cell is not None else f"עמודה_{idx+1}")

    data_rows = []
    last_seen = [None] * len(headers)

    for row in rows[header_row_index + 1:]:
        if not any(cell is not None for cell in row):
            continue

        current = []
        for i, cell in enumerate(row):
            value = clean_value(cell)
            if value is None:
                current.append(last_seen[i])
            else:
                current.append(value)
                last_seen[i] = value

        data_rows.append({headers[i]: current[i] for i in range(len(headers))})

    return data_rows

for file_key, file_path in INPUT_FILES.items():
    wb = load_workbook(file_path, data_only=True)
    all_rows = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = extract_sheet_rows(ws)

        for row in rows:
            row["sheet_name"] = sheet_name

        all_rows.extend(rows)

    output_file = OUTPUT_DIR / f"{file_key}.json"
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(all_rows, f, ensure_ascii=False, indent=2)

    print(f"created: {output_file}")