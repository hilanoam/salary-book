import json
from openpyxl import load_workbook
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent

file_path = BASE_DIR / "חוברת1.xlsx"
output_path = BASE_DIR / "public" / "data" / "professions.json"
output_path.parent.mkdir(parents=True, exist_ok=True)

print("מחפש קובץ כאן:", file_path)

if not file_path.exists():
    raise FileNotFoundError(f"הקובץ לא נמצא: {file_path}")

wb = load_workbook(file_path, data_only=True)
ws = wb.active

rows = list(ws.iter_rows(values_only=True))

headers = [
    str(h).strip() if h else f"col_{i+1}"
    for i, h in enumerate(rows[0])
]

data = []

for row in rows[1:]:
    if not any(row):
        continue

    record = {}

    for i, h in enumerate(headers):
        record[h] = row[i] if i < len(row) else None

    data.append(record)

with open(output_path, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"✔ professions.json נוצר עם {len(data)} שורות")
print("נשמר כאן:", output_path)