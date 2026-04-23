import json
from pathlib import Path
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent.parent
FILE = BASE_DIR / "נגדים.xlsx"
OUTPUT = BASE_DIR / "public" / "data" / "sergeants.json"

def clean(val):
    if val is None:
        return None
    if isinstance(val, float):
        return round(val, 2)
    return val

def normalize_headers(row):
    headers = []
    for i, cell in enumerate(row):
        if cell is None:
            headers.append(f"col_{i+1}")
        else:
            headers.append(str(cell).strip())
    return headers

def extract_group_number(sheet_name):
    if sheet_name.startswith("קבוצת תמריץ"):
        return sheet_name.replace("קבוצת תמריץ", "").strip()
    return None

def parse_sheet(ws, group_number):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(f"גיליון ריק: {ws.title}")
        return []

    headers = normalize_headers(rows[0])
    print(f"כותרות בגיליון {ws.title}: {headers}")

    data = []

    for row in rows[1:]:
        values = [clean(x) for x in row[:len(headers)]]

        if not any(v is not None for v in values):
            continue

        record = {"קבוצת תמריץ": group_number}

        for i, header in enumerate(headers):
            record[header] = values[i] if i < len(values) else None

        if not record.get("דרגת שכר"):
            continue

        data.append(record)

    print(f"נמצאו {len(data)} שורות בגיליון {ws.title}")
    return data

def main():
    print(f"מחפש קובץ כאן: {FILE}")

    if not FILE.exists():
        raise FileNotFoundError(f"הקובץ לא נמצא: {FILE}")

    wb = load_workbook(FILE, data_only=True)
    print(f"גיליונות שנמצאו: {wb.sheetnames}")

    all_rows = []

    for sheet_name in wb.sheetnames:
        group_number = extract_group_number(sheet_name)
        if not group_number:
            print(f"מדלג על גיליון: {sheet_name}")
            continue

        ws = wb[sheet_name]
        group_rows = parse_sheet(ws, group_number)
        all_rows.extend(group_rows)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(all_rows, f, ensure_ascii=False, indent=2)

    print(f"\nנוצר קובץ: {OUTPUT}")
    print(f"סה\"כ שורות: {len(all_rows)}")

if __name__ == "__main__":
    main()