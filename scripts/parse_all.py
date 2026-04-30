import json
from pathlib import Path
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR = BASE_DIR / "public" / "data"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

FILES = {
    "inspectors": BASE_DIR / "מפקח.xlsx",
    "lawyers": BASE_DIR / "משפטנים.xlsx",
    "sergeants": BASE_DIR / "נגדים.xlsx",
    "captain": BASE_DIR / "פקד.xlsx",
    "major": BASE_DIR / "רפק.xlsx",
}

def clean(val):
    if val is None:
        return None
    if isinstance(val, float):
        return round(val, 2)
    return val

def normalize_headers(row):
    return [str(h).strip() if h else f"col_{i}" for i, h in enumerate(row)]

def read_simple(ws, header_row=1, fixed_fields=None):
    rows = list(ws.iter_rows(values_only=True))
    headers = normalize_headers(rows[header_row - 1])

    data = []
    for row in rows[header_row:]:
        values = [clean(x) for x in row[:len(headers)]]

        if not any(values):
            continue

        record = {}

        if fixed_fields:
            record.update(fixed_fields)

        for i in range(len(headers)):
            record[headers[i]] = values[i] if i < len(values) else None

        if not record.get("דרגת שכר") and not record.get("סה\"כ משכורת ברוטו"):
            continue

        data.append(record)

    return data
# נגדים (כמו שעשינו)
def read_sergeants():
    wb = load_workbook(FILES["sergeants"], data_only=True)
    all_rows = []

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("קבוצת תמריץ"):
            continue

        group = sheet_name.replace("קבוצת תמריץ", "").strip()
        ws = wb[sheet_name]

        rows = list(ws.iter_rows(values_only=True))
        headers = normalize_headers(rows[0])

        for row in rows[1:]:
            values = [clean(x) for x in row[:len(headers)]]

            if not any(values):
                continue

            record = {"קבוצת תמריץ": group}
            for i, h in enumerate(headers):
                record[h] = values[i]

            if not record.get("דרגת שכר"):
                continue

            all_rows.append(record)

    return all_rows

def export(name, data):
    path = OUTPUT_DIR / f"{name}.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"✔ {name}.json → {len(data)} rows")

def main():
    wb = load_workbook(FILES["inspectors"], data_only=True)
    export(
        "inspectors",
        read_simple(wb.active, 1, {"דרגת שכר": "מפקח"})
    )

    wb = load_workbook(FILES["lawyers"], data_only=True)
    export(
        "lawyers",
        read_simple(wb.active, 1, {"דרגת שכר": "משפטנים"})
    )

    wb = load_workbook(FILES["captain"], data_only=True)
    export(
        "captain",
        read_simple(wb.active, 1, {"דרגת שכר": "פקד"})
    )

    wb = load_workbook(FILES["major"], data_only=True)
    export(
        "major",
        read_simple(wb.active, 1, {"דרגת שכר": "רפ\"ק"})
    )

    export("sergeants", read_sergeants())


if __name__ == "__main__":
    main()